def pagina_FIIs():
    from main import menu_principal
    import streamlit as st
    import pandas as pd
    import requests
    from datetime import date, timedelta
    from bs4 import BeautifulSoup
    import os
    import re
    from io import BytesIO
    import json
    import base64

    if st.button("⬅️ Voltar ao menu", key="btn_voltar_menu_fiis"):
        st.session_state.pagina = "menu"
        
    if st.session_state.pagina == "menu":
        menu_principal()
        st.rerun()

    # --- Persistência de Favoritos ---
    CAMINHO_FAVORITOS = "favoritos.json"

    def carregar_favoritos():
        if os.path.exists(CAMINHO_FAVORITOS):
            with open(CAMINHO_FAVORITOS, "r") as f:
                return json.load(f)
        return []

    def salvar_favoritos(lista):
        with open(CAMINHO_FAVORITOS, "w") as f:
            json.dump(lista, f)

    def buscar_noticias(fii, data_inicial_str, data_final_str):
        url = "https://sistemasweb.b3.com.br/PlantaoNoticias/Noticias/ListarTitulosNoticias"
        params = {
            "agencia": "18",
            "palavra": re.sub(r'\d+', '', fii),
            "dataInicial": data_inicial_str,
            "dataFinal": data_final_str
        }
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": "https://sistemasweb.b3.com.br/PlantaoNoticias/Noticias"
        }
        resultados = []
        dados_dividendos = {
            "Fundo": fii.upper(),
            "Data-Base": None,
            "Data Pagamento": None,
            "Último Dividendo (R$)": None,
            "Link Relatório": None
        }
        try:
            response = requests.get(url, params=params, headers=headers, timeout=15)
            response.raise_for_status()
            dados = response.json()

            for n in dados:
                if "NwsMsg" not in n:
                    continue
                nws_msg = n["NwsMsg"]
                nws_cod = nws_msg["id"]
                nws_data = nws_msg["dateTime"]
                titulo = nws_msg["headline"]

                noticia_info = {
                    "Fundo": fii,
                    "Data": nws_data,
                    "Título": titulo,
                    "Cod": nws_cod,
                    "Link": f"https://sistemasweb.b3.com.br/PlantaoNoticias/Noticias/Detail?idNoticia={nws_cod}&agencia=18&dataNoticia={nws_data}"
                }
                resultados.append(noticia_info)

                # Extrai dividendos só se a notícia for de cotistas E se ainda não pegou dividendo deste FII
                if "cotistas" in titulo.lower() and dados_dividendos["Data-Base"] is None:
                    try:
                        r_det = requests.get(noticia_info["Link"], headers=headers, timeout=10)
                        soup = BeautifulSoup(r_det.text, "html.parser")
                        conteudo_pre = soup.find("pre", id="conteudoDetalhe")
                        if conteudo_pre:
                            pre_text = conteudo_pre.get_text()
                            match = re.search(r"https://fnet\.bmfbovespa\.com\.br/fnet/publico/visualizarDocumento\?[^ )\n]+", pre_text)
                            if match:
                                url_doc = match.group(0).replace("&amp;", "&")
                                r_doc = requests.get(url_doc, headers=headers, timeout=15)
                                soup_doc = BeautifulSoup(r_doc.text, "html.parser")
                                iframe = soup_doc.find("iframe")
                                if iframe and iframe.get("src"):
                                    base_url = "https://fnet.bmfbovespa.com.br/fnet/publico/"
                                    iframe_src = iframe["src"]
                                    iframe_url = iframe_src if iframe_src.startswith("http") else base_url + iframe_src.lstrip("/")
                                    r_iframe = requests.get(iframe_url, headers=headers, timeout=15)
                                    try:
                                        conteudo = r_iframe.content.decode("utf-8")
                                        decoded = base64.b64decode(conteudo).decode("utf-8")
                                        soup_iframe = BeautifulSoup(decoded, "html.parser")
                                        tabela = [span.text.strip() for span in soup_iframe.find_all("span", class_="dado-valores") if span.text.strip()]
                                        tabela = list(dict.fromkeys(tabela))
                     
                                        if len (tabela ) < 7 or len (tabela) >= 11:
                                            dados_dividendos["Data-Base"] = tabela[0]
                                            dados_dividendos["Data Pagamento"] = tabela[2]
                                            dados_dividendos["Último Dividendo (R$)"] = tabela[1]
                                            dados_dividendos["Link Relatório"] = url_doc
                                        else:
                                            dados_dividendos["Data-Base"] = tabela[0]
                                            dados_dividendos["Data Pagamento"] = tabela[3]
                                            dados_dividendos["Último Dividendo (R$)"] = tabela[2]
                                            dados_dividendos["Link Relatório"] = url_doc 
                                    except Exception:
                                        pass
                    except Exception as e:
                        st.warning(f"Falha ao buscar dividendos: {e}")
        except Exception as e:
            st.error(f"Erro ao buscar notícias para {fii}: {str(e)}")
        return resultados, dados_dividendos

    # --- Layout ---
    st.set_page_config("Analisador FIIs", layout="wide")
    st.title("🔍 Analisador de FIIs - Notícias e Dividendos")

    # --- Lista FIIs ---
    todos_fiis =  [
        'HGRU11', 'KEVE11', 'BTAL11', 'XPML11', 'JSRE11', 'TVRI11',
        'RECR11', 'KNRI11', 'RBRF11', 'MXRF11', 'RZTR11', 'XPCI11', 'BRCO11', 'HGRE11', 'VGIR11',
        'XPLG11', 'CPTS11', 'ALZR11', 'KFOF11', 'IRDM11', 'RBRR11', 'KNHF11', 'VISC11', 'HGLG11', 'VIUR11',
        'BTLG11',  'KNHY11', 'MCCI11', 'RBRY11', 'KNSC11', 'PVBI11', 'HSML11', 'KNUQ11', 'KNCR11',
        'LVBI11', 'FATN11', 'GGRC11', 'KORE11', 'TRXF11', 'HGCR11', 'HGFF11', 'VINO11', 'TGAR11', 'KNIP11',
        'RBVA11', 'VILG11', 'VCJR11'
    ]

    def init_session():
        if "favoritos" not in st.session_state:
            st.session_state.favoritos = carregar_favoritos()
        if "filtro" not in st.session_state:
            st.session_state.filtro = ""
        for fii in todos_fiis:
            if f"chk_{fii}" not in st.session_state:
                st.session_state[f"chk_{fii}"] = False

    init_session()

    hoje = date.today()
    min_date = hoje - timedelta(days=30)
    data_inicial = st.date_input(
        "Selecione a data inicial para as notícias",
        min_value=min_date,
        max_value=hoje,
        value=hoje - timedelta(days=3),
        format="DD/MM/YYYY"
    )
    data_final = hoje
    data_inicial_str = data_inicial.strftime("%Y-%m-%d")
    data_final_str = data_final.strftime("%Y-%m-%d")

    filtro = st.text_input("🔎 Filtrar FIIs por nome ou ticker:", value=st.session_state.get("filtro", ""), key="filtro_input")
    fiis_filtrados = [f for f in todos_fiis if st.session_state["filtro_input"].upper() in f.upper()]
    fiis_filtrados.sort()

    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        if st.button("Selecionar Todos"):
            for fii in fiis_filtrados:
                st.session_state[f"chk_{fii}"] = True
    with col2:
        if st.button("Limpar Seleção"):
            for fii in fiis_filtrados:
                st.session_state[f"chk_{fii}"] = False
    with col3:
        if st.button("Selecionar Favoritos"):
            for fii in fiis_filtrados:
                st.session_state[f"chk_{fii}"] = fii in st.session_state.favoritos

    st.markdown("### 🎯 Selecione os FIIs a analisar:")
    cols = st.columns(3)
    for i, fii in enumerate(fiis_filtrados):
        with cols[i % 3]:
            col_star, col_chk = st.columns([0.15, 0.85])
            with col_star:
                icone = "⭐" if fii in st.session_state.favoritos else "☆"
                if st.button(icone, key=f"fav_{fii}"):
                    if fii in st.session_state.favoritos:
                        st.session_state.favoritos.remove(fii)
                    else:
                        st.session_state.favoritos.append(fii)
                    salvar_favoritos(st.session_state.favoritos)
                    st.rerun()
            with col_chk:
                st.checkbox(
                    fii,
                    value=st.session_state[f"chk_{fii}"],
                    key=f"chk_{fii}"
                )

    selecionados = [fii for fii in todos_fiis if st.session_state.get(f"chk_{fii}", False)]

    if st.button("🚀 Analisar Selecionados") and selecionados:
        with st.spinner("🔄 Coletando dados..."):
            todas_noticias = []
            dividendos_atuais = []

            for fii in selecionados:
                noticias_fii, dividendo = buscar_noticias(fii, data_inicial_str, data_final_str)
                todas_noticias.extend(noticias_fii)
                dividendos_atuais.append(dividendo)

            # Notas: df_noticias NÃO tem colunas de dividendos!
            df_noticias = pd.DataFrame(todas_noticias)[["Fundo", "Data", "Título", "Cod", "Link"]]
            df_atuais = pd.DataFrame(dividendos_atuais)

            # Histórico dividendos
            historico_path = "historico_dividendos.csv"
            if os.path.exists(historico_path):
                df_anterior = pd.read_csv(historico_path)
            else:
                df_anterior = pd.DataFrame(columns=["Fundo", "Data-Base", "Data Pagamento", "Último Dividendo (R$)"])

            def comparar(row):
                anterior = df_anterior[df_anterior["Fundo"] == row["Fundo"]]
                if anterior.empty:
                    return "NOVO"
                elif row["Data-Base"] != anterior["Data-Base"].values[0]:
                    return "NOVA DATA-BASE"
                elif row["Último Dividendo (R$)"] != anterior["Último Dividendo (R$)"].values[0]:
                    return "VALOR ALTERADO"
                return "IGUAL"

            if not df_atuais.empty:
                df_atuais["Status"] = df_atuais.apply(comparar, axis=1)
                df_atuais.to_csv(historico_path, index=False)

            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border
            from openpyxl.utils import get_column_letter

            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_noticias.to_excel(writer, sheet_name="Notícias", index=False)
                df_atuais.to_excel(writer, sheet_name="Dividendos", index=False)
            buffer.seek(0)

            wb = load_workbook(buffer)
            for ws in wb.worksheets:
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill("solid", fgColor="4472C4")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border()
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border()
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value is not None:
                            value_length = len(str(cell.value))
                            if value_length > max_length:
                                max_length = value_length
                    ws.column_dimensions[col_letter].width = max(15, min(max_length + 2, 80))
                ws.sheet_view.showGridLines = False

            output = BytesIO()
            wb.save(output)
            output.seek(0)

        st.markdown("### 📝 Notícias Recentes")
        with st.expander("Mostrar/Ocultar", expanded=True):
            if df_noticias.empty:
                st.info("Nenhuma notícia encontrada.")
            else:
                st.dataframe(df_noticias
                    .rename(columns={
                        "Fundo": "Fundo",
                        "Data": "Data",
                        "Título": "Título"
                    })
                    .style.set_properties(**{'text-align': 'left'})
                    .set_table_styles([{
                        'selector': 'th',
                        'props': [('background-color', '#4472C4'),
                                ('color', 'white'),
                                ('font-weight', 'bold')]
                    }]))

        st.markdown("---")
        st.markdown("### 🔴 Comparativo de Dividendos")
        with st.expander("Mostrar/Ocultar", expanded=True):
            st.dataframe(df_atuais[["Fundo", "Data-Base", "Data Pagamento", "Último Dividendo (R$)", "Status", "Link Relatório"]]
                    .rename(columns={
                        "Fundo": "Fundo",
                        "Data-Base": "Data-Base",
                        "Pagamento": "Pagamento",
                        "Último Dividendo (R$)": "Último Dividendo (R$)",
                        "Status": "Status"
                    })
                    .style.set_properties(**{'text-align': 'left'})
                    .set_table_styles([{
                        'selector': 'th',
                        'props': [('background-color', '#4472C4'),
                                ('color', 'white'),
                                ('font-weight', 'bold')]
                    }]))
            alterados = df_atuais[df_atuais["Status"] == "VALOR ALTERADO"]
            if not alterados.empty:
                st.warning("🔴 Fundos com alteração no dividendo:")
                for _, row in alterados.iterrows():
                    fundo = row["Fundo"]
                    atual = row["Último Dividendo (R$)"]
                    anterior = df_anterior[df_anterior["Fundo"] == fundo]["Último Dividendo (R$)"].values[0]
    

        st.markdown("---")
        st.markdown("### 📝 Baixar Excel com Dados")
        st.download_button(
            "⬇️ Baixar Planilha Completa",
            data=output,
            file_name="FIIs_noticias_dividendos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )