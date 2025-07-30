import pandas as pd
from datetime import datetime, timedelta
import urllib.request
import json
from openpyxl import load_workbook

# Carregamento dos dados importados
from lista_fundos_analisados import df_precos, fundos_raw, nomes_fundos_limpos, arquivo

# Tratamento dos preços
precos = df_precos.iloc[6:, 1:2+len(fundos_raw)].copy()
precos.columns = ["Data"] + nomes_fundos_limpos
precos["Data"] = pd.to_datetime(precos["Data"], format="%d/%m/%Y", errors="coerce")
precos = precos.dropna(subset=["Data"])
precos = precos.sort_values("Data").reset_index(drop=True)
ontem = datetime.now().date() - timedelta(days=1)
precos = precos[precos["Data"].dt.date <= ontem].reset_index(drop=True)

# CDI - Dados do Banco Central
data_inicio = '02/01/2017'
data_fim = datetime.today().strftime('%d/%m/%Y')
url = f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.12/dados?formato=json&dataInicial={data_inicio}&dataFinal={data_fim}"

with urllib.request.urlopen(url) as response:
    data = response.read()
    data_json = json.loads(data)
    df_cdi = pd.DataFrame(data_json)
    df_cdi['data'] = pd.to_datetime(df_cdi['data'], format='%d/%m/%Y')
    df_cdi['valor'] = pd.to_numeric(df_cdi['valor'])
    df_cdi.columns = ['Data', 'Fator diário do CDI']
    df_cdi['Fator diário do CDI'] = 1 + (df_cdi['Fator diário do CDI'] / 100)
    df_cdi["Taxa anual"] = df_cdi["Fator diário do CDI"]**252 - 1

# Merge preços e CDI
df_merged = pd.merge(precos, df_cdi, on="Data", how="inner")

df_cdi_mensal = df_cdi.copy()
df_cdi_mensal["MesAno"] = df_cdi_mensal["Data"].dt.to_period("M").dt.to_timestamp()
df_cdi_mensal = df_cdi_mensal.groupby("MesAno")["Taxa anual"].last().reset_index()
df_cdi_mensal = df_cdi_mensal.set_index("MesAno")
serie_cdi = df_cdi_mensal["Taxa anual"]


# Cálculo dos alfas
alfas = pd.DataFrame()
alfas["Data"] = df_merged["Data"]
for fundo in precos.columns[1:]:
    preco_hoje = df_merged[fundo]
    preco_ontem = df_merged[fundo].shift(1)
    fator_cdi = df_merged['Fator diário do CDI']
    alfa = (preco_hoje / (preco_ontem * fator_cdi)) - 1
    alfas[fundo] = alfa

# Limpeza e filtro dos alfas
ontem = alfas["Data"].max()
alfas = alfas[alfas["Data"] < ontem].reset_index(drop=True)
mask = (alfas['Data'] >= '2020-03-01') & (alfas['Data'] <= '2021-12-01')
alfas = alfas[~mask]

# Leitura do Excel
abas = pd.read_excel(arquivo, sheet_name=None)

# Preparação do DataFrame de DY diário
df_dy_diario = pd.DataFrame()
df_dy_diario["Data"] = precos["Data"]

# Leitura do arquivo Excel com preços e dividendos
wb = load_workbook(arquivo, data_only=True)

for nome_fundo in precos.columns:
    if nome_fundo in ["Data", "Preços", "Variaveis"] or nome_fundo not in abas:
        continue
    try:
        aba = abas[nome_fundo]
        df_div = aba.iloc[1:, [1, 2]].copy()
        df_div.columns = ["Data", "Dividendo"]
        df_div["Data"] = pd.to_datetime(df_div["Data"], errors="coerce")
        df_div["Dividendo"] = pd.to_numeric(df_div["Dividendo"], errors="coerce")
        df_div = df_div.dropna(subset=["Data", "Dividendo"])
        if df_div.empty:
            continue
        df_precos_fundo = precos[["Data", nome_fundo]].copy().rename(columns={nome_fundo: "Preco"})
        df_precos_fundo = df_precos_fundo.dropna(subset=["Preco"])
        df_div_diario = pd.DataFrame({"Data": df_precos_fundo["Data"]})
        df_div_diario = df_div_diario.merge(df_div, on="Data", how="left")
        df_div_diario["Dividendo"] = df_div_diario["Dividendo"].ffill()
        df_merge = df_div_diario.merge(df_precos_fundo, on="Data", how="inner")
        df_merge["DY_diario"] = (df_merge["Dividendo"] * 12) / df_merge["Preco"]

        if nome_fundo in wb.sheetnames:
            ws = wb[nome_fundo]
            datas_desdobramento = []
            for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
                cell = row[0].value
                if isinstance(cell, (datetime, pd.Timestamp)):
                    datas_desdobramento.append(pd.to_datetime(cell))
            for data_desdobramento in datas_desdobramento:
                df_merge.loc[df_merge["Data"] < data_desdobramento, "DY_diario"] /= 10

        df_dy_diario = df_dy_diario.merge(
            df_merge[["Data", "DY_diario"]].rename(columns={"DY_diario": nome_fundo}),
            on="Data", how="left"
        )
    except Exception as e:
        print(f"[{nome_fundo}] Erro: {e}")

# Construção do df_dy_mensal usando o último DY diário de cada mês
df_dy_diario["Data"] = pd.to_datetime(df_dy_diario["Data"], errors="coerce")
df_dy_diario["MesAno"] = df_dy_diario["Data"].dt.to_period("M").dt.to_timestamp()
df_dy_mensal = df_dy_diario.drop(columns=["Data"]).groupby("MesAno").last()
df_dy_mensal = df_dy_mensal.sort_index().round(4)