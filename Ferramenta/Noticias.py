def pagina_FIIs():
    import streamlit as st
    import pandas as pd
    import requests
    from datetime import date, timedelta
    from bs4 import BeautifulSoup
    import os
    import re
    from io import BytesIO
    import json
    import base64
    from pathlib import Path

    st.write("📁 Diretório de execução atual:", os.getcwd())
    st.write("📂 Diretório onde está o script:", Path(__file__).parent)
    st.write("📄 Caminho completo do arquivo:", Path(__file__).parent / "historico_dividendos_formatado_string_virgula.csv")
    st.write("✅ Arquivo existe?", (Path(__file__).parent / "historico_dividendos_formatado_string_virgula.csv").exists())

    # --- Persistência de Favoritos ---
    CAMINHO_FAVORITOS = "favoritos.json"

    def carregar_favoritos():
        if os.path.exists(CAMINHO_FAVORITOS):
            with open(CAMINHO_FAVORITOS, "r") as f:
                return json.load(f)
        return []

    def salvar_favoritos(lista):
        with open(CAMINHO_FAVORITOS, "w") as f:
            json.dump(lista, f)

    def buscar_noticias(fii, data_inicial_str, data_final_str):
        url = "https://sistemasweb.b3.com.br/PlantaoNoticias/Noticias/ListarTitulosNoticias"
        params = {
            "agencia": "18",
            "palavra": re.sub(r'\d+', '', fii),
            "dataInicial": data_inicial_str,
            "dataFinal": data_final_str
        }
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": "https://sistemasweb.b3.com.br/PlantaoNoticias/Noticias"
        }
        resultados = []
        dados_dividendos = {
            "Fundo": fii.upper(),
            "Data-Base": None,
            "Data Pagamento": None,
            "Último Dividendo (R$)": None,
            "Link Relatório": None
        }
        try:
            response = requests.get(url, params=params, headers=headers, timeout=30)
            response.raise_for_status()
            dados = response.json()

            for n in dados:
                if "NwsMsg" not in n:
                    continue
                nws_msg = n["NwsMsg"]
                nws_cod = nws_msg["id"]
                nws_data = nws_msg["dateTime"]
                titulo = nws_msg["headline"]

                noticia_info = {
                    "Fundo": fii,
                    "Data": nws_data,
                    "Título": titulo,
                    "Cod": nws_cod,
                    "Link": f"https://sistemasweb.b3.com.br/PlantaoNoticias/Noticias/Detail?idNoticia={nws_cod}&agencia=18&dataNoticia={nws_data}"
                }
                resultados.append(noticia_info)

                # Extrai dividendos só se a notícia for de cotistas E se ainda não pegou dividendo deste FII
                if "cotistas" in titulo.lower() and dados_dividendos["Data-Base"] is None:
                    try:
                        r_det = requests.get(noticia_info["Link"], headers=headers, timeout=10)
                        soup = BeautifulSoup(r_det.text, "html.parser")
                        conteudo_pre = soup.find("pre", id="conteudoDetalhe")
                        if conteudo_pre:
                            pre_text = conteudo_pre.get_text()
                            match = re.search(r"https://fnet\.bmfbovespa\.com\.br/fnet/publico/visualizarDocumento\?[^ )\n]+", pre_text)
                            if match:
                                url_doc = match.group(0).replace("&amp;", "&")
                                r_doc = requests.get(url_doc, headers=headers, timeout=15)
                                soup_doc = BeautifulSoup(r_doc.text, "html.parser")
                                iframe = soup_doc.find("iframe")
                                if iframe and iframe.get("src"):
                                    base_url = "https://fnet.bmfbovespa.com.br/fnet/publico/"
                                    iframe_src = iframe["src"]
                                    iframe_url = iframe_src if iframe_src.startswith("http") else base_url + iframe_src.lstrip("/")
                                    r_iframe = requests.get(iframe_url, headers=headers, timeout=15)
                                    try:
                                        conteudo = r_iframe.content.decode("utf-8")
                                        decoded = base64.b64decode(conteudo).decode("utf-8")
                                        soup_iframe = BeautifulSoup(decoded, "html.parser")
                                        tabela = [span.text.strip() for span in soup_iframe.find_all("span", class_="dado-valores") if span.text.strip()]
                                        tabela = list(dict.fromkeys(tabela))
                     
                                        if len (tabela ) < 7 or len (tabela) >= 11:
                                            dados_dividendos["Data-Base"] = tabela[0]
                                            dados_dividendos["Data Pagamento"] = tabela[2]
                                            dados_dividendos["Último Dividendo (R$)"] = tabela[1]
                                            dados_dividendos["Link Relatório"] = url_doc
                                        else:
                                            dados_dividendos["Data-Base"] = tabela[0]
                                            dados_dividendos["Data Pagamento"] = tabela[3]
                                            dados_dividendos["Último Dividendo (R$)"] = tabela[2]
                                            dados_dividendos["Link Relatório"] = url_doc 
                                    except Exception:
                                        pass
                    except Exception as e:
                        st.warning(f"Falha ao buscar dividendos: {e}")
        except Exception as e:
            st.error(f"Erro ao buscar notícias para {fii}: {str(e)}")
        return resultados, dados_dividendos

    # --- Layout ---
    st.set_page_config("Analisador FIIs", layout="wide")
    st.title("🔍 Analisador de FIIs - Notícias e Dividendos")

    # --- Lista FIIs ---
    from lista_fundos_analisados import nomes_fundos_limpos

    def init_session():
        if "favoritos" not in st.session_state:
            st.session_state.favoritos = carregar_favoritos()
        if "filtro" not in st.session_state:
            st.session_state.filtro = ""
        for fii in nomes_fundos_limpos:
            if f"chk_{fii}" not in st.session_state:
                st.session_state[f"chk_{fii}"] = False

    init_session()

    hoje = date.today()
    min_date = hoje - timedelta(days=30)
    data_inicial = st.date_input(
        "Selecione a data inicial para as notícias",
        min_value=min_date,
        max_value=hoje,
        value=hoje - timedelta(days=3),
        format="DD/MM/YYYY"
    )
    data_final = hoje
    data_inicial_str = data_inicial.strftime("%Y-%m-%d")
    data_final_str = data_final.strftime("%Y-%m-%d")

    filtro = st.text_input("🔎 Filtrar FIIs por nome ou ticker:", value=st.session_state.get("filtro", ""), key="filtro_input")
    fiis_filtrados = [f for f in nomes_fundos_limpos if st.session_state["filtro_input"].upper() in f.upper()]
    fiis_filtrados.sort()

    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        if st.button("Selecionar Todos"):
            for fii in fiis_filtrados:
                st.session_state[f"chk_{fii}"] = True
    with col2:
        if st.button("Limpar Seleção"):
            for fii in fiis_filtrados:
                st.session_state[f"chk_{fii}"] = False
    with col3:
        if st.button("Selecionar Favoritos"):
            for fii in fiis_filtrados:
                st.session_state[f"chk_{fii}"] = fii in st.session_state.favoritos

    st.markdown("### 🎯 Selecione os FIIs a analisar:")
    cols = st.columns(3)
    for i, fii in enumerate(fiis_filtrados):
        with cols[i % 3]:
            col_star, col_chk = st.columns([0.15, 0.85])
            with col_star:
                icone = "⭐" if fii in st.session_state.favoritos else "☆"
                if st.button(icone, key=f"fav_{fii}"):
                    if fii in st.session_state.favoritos:
                        st.session_state.favoritos.remove(fii)
                    else:
                        st.session_state.favoritos.append(fii)
                    salvar_favoritos(st.session_state.favoritos)
                    st.rerun()
            with col_chk:
                st.checkbox(
                    fii,
                    value=st.session_state[f"chk_{fii}"],
                    key=f"chk_{fii}"
                )

    selecionados = [fii for fii in nomes_fundos_limpos if st.session_state.get(f"chk_{fii}", False)]

    if st.button("🚀 Analisar Selecionados") and selecionados:
        with st.spinner("🔄 Coletando dados..."):
            todas_noticias = []
            dividendos_atuais = []

            progress_bar = st.progress(0)
            status_text = st.empty()

            total = len(selecionados)
            for idx, fii in enumerate(selecionados, 1):
                status_text.markdown(f"🔎 Coletando <b>{idx}</b> de <b>{total}</b> ({fii})...", unsafe_allow_html=True)
                noticias_fii, dividendo = buscar_noticias(fii, data_inicial_str, data_final_str)
                todas_noticias.extend(noticias_fii)
                dividendos_atuais.append(dividendo)
                progress_bar.progress(idx / total)

            status_text.markdown(f"✅ Coleta finalizada: {total} de {total} FIIs processados.", unsafe_allow_html=True)
            progress_bar.empty()

            # Notas: df_noticias NÃO tem colunas de dividendos!
            df_noticias = pd.DataFrame(todas_noticias)[["Fundo", "Data", "Título", "Cod", "Link"]]
            df_atuais = pd.DataFrame(dividendos_atuais)

            from pathlib import Path
            import pandas as pd
            import streamlit as st

            historico_path = Path(__file__).parent / "historico_dividendos_formatado_string_virgula.csv"
            colunas_hist = [
                "Fundo", "Último Data-Base", "Último Pagamento", "Último Dividendo (R$)",
                "Anterior Data-Base", "Anterior Pagamento", "Anterior Dividendo (R$)", "Status", "Link Relatório"
            ]

            # Debug visual
            st.write("📄 Arquivo esperado:", historico_path)
            st.write("✅ Arquivo existe?", historico_path.exists())

            # Leitura segura
            if historico_path.exists():
                st.success("✅ Entrou no IF: arquivo encontrado e será lido.")
                df_anterior = pd.read_csv(historico_path)
            else:
                st.error("❌ Arquivo não encontrado. Criando DataFrame vazio.")
                df_anterior = pd.DataFrame(columns=colunas_hist)


            def atualizar_historico(df_novo, df_hist):
                fundos_existentes = set(df_hist["Fundo"])
                fundos_processados = set(df_novo["Fundo"])
                todos_fundos = fundos_existentes.union(fundos_processados)

                registros_atualizados = []

                for fundo in todos_fundos:
                    linha_nova = df_novo[df_novo["Fundo"] == fundo]
                    linha_antiga = df_hist[df_hist["Fundo"] == fundo]

                    if linha_nova.empty:
                        # Caso 5: fundo não foi processado nessa rodada
                        if not linha_antiga.empty:
                            registros_atualizados.append(linha_antiga.iloc[0].to_dict())
                        continue

                    novo_db = linha_nova["Data-Base"].values[0]
                    novo_pg = linha_nova["Data Pagamento"].values[0]
                    novo_valor = linha_nova["Último Dividendo (R$)"].values[0]
                    link = linha_nova["Link Relatório"].values[0]

                    if linha_antiga.empty:
                        # Caso 1: fundo novo
                        registros_atualizados.append({
                            "Fundo": fundo,
                            "Último Data-Base": novo_db,
                            "Último Pagamento": novo_pg,
                            "Último Dividendo (R$)": novo_valor,
                            "Anterior Data-Base": None,
                            "Anterior Pagamento": None,
                            "Anterior Dividendo (R$)": None,
                            "Link Relatório": link
                        })
                    else:
                        # Fundo já existia, comparar para ver se atualiza
                        ult_db = linha_antiga["Último Data-Base"].values[0]
                        ult_pg = linha_antiga["Último Pagamento"].values[0]
                        ult_valor = linha_antiga["Último Dividendo (R$)"].values[0]
                        ant_db = linha_antiga["Anterior Data-Base"].values[0]
                        ant_pg = linha_antiga["Anterior Pagamento"].values[0]
                        ant_valor = linha_antiga["Anterior Dividendo (R$)"].values[0]
                        link_antigo = linha_antiga["Link Relatório"].values[0] if "Link Relatório" in linha_antiga else None

                        if pd.notnull(novo_db):
                            if (novo_db != ult_db) or (str(novo_valor).replace(",", ".") != str(ult_valor).replace(",", ".")):
                                # Casos 2 e 4: nova distribuição (data nova OU valor diferente)
                                registros_atualizados.append({
                                    "Fundo": fundo,
                                    "Último Data-Base": novo_db,
                                    "Último Pagamento": novo_pg,
                                    "Último Dividendo (R$)": novo_valor,
                                    "Anterior Data-Base": ult_db,
                                    "Anterior Pagamento": ult_pg,
                                    "Anterior Dividendo (R$)": ult_valor,
                                    "Link Relatório": link
                                })
                            else:
                                # Caso 3: nada mudou de fato
                                registros_atualizados.append({
                                    "Fundo": fundo,
                                    "Último Data-Base": ult_db,
                                    "Último Pagamento": ult_pg,
                                    "Último Dividendo (R$)": ult_valor,
                                    "Anterior Data-Base": ant_db,
                                    "Anterior Pagamento": ant_pg,
                                    "Anterior Dividendo (R$)": ant_valor,
                                    "Link Relatório": link_antigo
                                })
                        else:
                            # Caso 6: fundo foi processado mas sem dados novos
                            registros_atualizados.append({
                                "Fundo": fundo,
                                "Último Data-Base": ult_db,
                                "Último Pagamento": ult_pg,
                                "Último Dividendo (R$)": ult_valor,
                                "Anterior Data-Base": ant_db,
                                "Anterior Pagamento": ant_pg,
                                "Anterior Dividendo (R$)": ant_valor,
                                "Link Relatório": link_antigo
                            })

                return pd.DataFrame(registros_atualizados)



            if not df_atuais.empty:
                df_atuais_atualizado = atualizar_historico(df_atuais, df_anterior)
                df_atuais_atualizado.to_csv(historico_path, index=False)
            else:
                df_atuais_atualizado = df_anterior.copy()

            df_exibicao = df_atuais_atualizado[df_atuais_atualizado["Fundo"].isin(df_atuais["Fundo"])]


            # --- Exibição no Streamlit ---

            st.markdown("### 📝 Notícias Recentes")
            with st.expander("Mostrar/Ocultar", expanded=True):
                if df_noticias.empty:
                    st.info("Nenhuma notícia encontrada.")
                else:
                    st.dataframe(df_noticias
                        .rename(columns={
                            "Fundo": "Fundo",
                            "Data": "Data",
                            "Título": "Título"
                        })
                        .style.set_properties(**{'text-align': 'left'})
                        .set_table_styles([{
                            'selector': 'th',
                            'props': [('background-color', '#4472C4'),
                                    ('color', 'white'),
                                    ('font-weight', 'bold')]
                        }]))

            st.markdown("---")
            st.markdown("### 🔴 Comparativo de Dividendos")
            with st.expander("Mostrar/Ocultar", expanded=True):
                colunas_exibidas = [
                    "Fundo",
                    "Anterior Data-Base", "Anterior Pagamento", "Anterior Dividendo (R$)",
                    "Último Data-Base", "Último Pagamento", "Último Dividendo (R$)",
                    "Link Relatório"
                ]
                st.dataframe(df_exibicao[colunas_exibidas]
                    .style.set_properties(**{'text-align': 'left'})
                    .set_table_styles([{
                        'selector': 'th',
                        'props': [('background-color', '#4472C4'),
                                ('color', 'white'),
                                ('font-weight', 'bold')]
                    }])
                )

                # Conversão segura

                # Padroniza separadores decimais (vírgula para ponto) antes de converter
                div_ult_raw = df_exibicao["Último Dividendo (R$)"].astype(str).str.replace(",", ".")
                div_ant_raw = df_exibicao["Anterior Dividendo (R$)"].astype(str).str.replace(",", ".")

                # Converte para float com segurança
                div_ult = pd.to_numeric(div_ult_raw, errors='coerce')
                div_ant = pd.to_numeric(div_ant_raw, errors='coerce')

                # Detecta alterações reais
                diferenca = (div_ult != div_ant) & div_ult.notnull() & div_ant.notnull()
                alterados = df_exibicao[diferenca]


                if not alterados.empty:
                    st.warning("🔴 Fundos com alteração no valor do dividendo:")
                    
                    # Seleciona apenas colunas relevantes para exibir
                    colunas_alteradas = [
                        "Fundo",
                        "Anterior Data-Base", "Anterior Pagamento", "Anterior Dividendo (R$)",
                        "Último Data-Base", "Último Pagamento", "Último Dividendo (R$)",
                        "Link Relatório"
                    ]
                    
                    st.dataframe(
                        alterados[colunas_alteradas]
                        .style.set_properties(**{'text-align': 'left'})
                        .set_table_styles([{
                            'selector': 'th',
                            'props': [('background-color', '#FFD966'),  # amarelo claro
                                    ('color', 'black'),
                                    ('font-weight', 'bold')]
                        }])
                    )
                else:
                    st.warning("🔄 Nenhum fundo sofreu alteração no valor do dividendo.")
                
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df_noticias.to_excel(writer, sheet_name="Notícias", index=False)
                    df_exibicao.to_excel(writer, sheet_name="Dividendos", index=False)
                buffer.seek(0)

                from openpyxl import load_workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border
                from openpyxl.utils import get_column_letter

                # Fundo com alteração detectada
                fundos_alterados = set(alterados["Fundo"])

                wb = load_workbook(buffer)
                for ws in wb.worksheets:
                    header_font = Font(bold=True, color="FFFFFF", size=12)
                    header_fill = PatternFill("solid", fgColor="4472C4")
                    destaque_fill = PatternFill("solid", fgColor="C0E6F5")  # Azul claro para linhas alteradas

                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border()

                    for row in ws.iter_rows(min_row=2):
                        fundo_cell = row[0].value  # Primeira coluna deve ser "Fundo"
                        is_alterado = ws.title == "Dividendos" and fundo_cell in fundos_alterados
                        for cell in row:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border = Border()
                            if is_alterado:
                                cell.fill = destaque_fill

                    for col in ws.columns:
                        max_length = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            if cell.value is not None:
                                value_length = len(str(cell.value))
                                if value_length > max_length:
                                    max_length = value_length
                        ws.column_dimensions[col_letter].width = max(15, min(max_length + 2, 80))
                    ws.sheet_view.showGridLines = False

                output = BytesIO()
                wb.save(output)
                output.seek(0)

        st.markdown("---")
        st.markdown("### 📝 Baixar Excel com Dados")
        st.download_button(
            "⬇️ Baixar Planilha Completa",
            data=output,
            file_name="FIIs_noticias_dividendos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )